import os
import math
from matplotlib import pyplot
import numpy as np
import cv2
import torch
import torch.nn as nn
from torch.nn import functional
import openpyxl as excel


def ndarray_to_tensor(x, is_cuda=True, requires_grad=False, dtype=torch.float32):
    t = torch.tensor(x, dtype=dtype, requires_grad=requires_grad)
    if is_cuda:
        t = t.cuda()
    return t


def log(path, log, mode='a', line=None, is_print=True):
    # line int i: add line i-th line of existing text. None: add line at the end of existing text
    if line is not None:
        with open(path, 'r') as file:
            exist_text_list = file.readlines()

        if not isinstance(line, int):
            raise Exception('invalid line')

        # add new lines
        line_now = line
        for l in log:
            exist_text_list.insert(line_now, l + '\n')
            line_now += 1
            if is_print:
                print(l, flush=True)

        # write to file
        with open(path, 'w') as file:
            file.writelines(exist_text_list)
    else:
        with open(path, mode) as file:
            for l in log:
                file.write(l + '\n')
                if is_print:
                    print(l, flush=True)
    log[:] = []


def is_only_one_bool_is_true(*flag):
    count = 0
    for f in flag:
        if not isinstance(f, bool):
            raise Exception('not supported type')
        elif f:
            count += 1
    if count == 1:
        return True
    else:
        return False


def show_matrix(matrix_list):
    # matrix a list of 2 dimension numpy.ndarray
    for matrix in matrix_list:
        pyplot.figure()
        pyplot.imshow(matrix)
    pyplot.show()
    return


def compare_result(result_dict, best_result_dict, key_value, reverse=False):
    # Returns a dict of dictionaries containing the best error specified by the key_value
    # result_dict: dict
    # best_result_dict: dict
    # key_value: string
    # reverse: bool, False: return a result with smaller key value, True: return a result with larger key value
    final_result_dict = dict()
    for data_name in result_dict:
        result = result_dict[data_name]
        best_result = best_result_dict[data_name]

        if reverse:
            if result[key_value] > best_result[key_value]:
                final_result_dict[data_name] = result
            else:
                final_result_dict[data_name] = best_result
        else:
            if result[key_value] < best_result[key_value]:
                final_result_dict[data_name] = result
            else:
                final_result_dict[data_name] = best_result
    return final_result_dict


def compare_mae(correct_cent_list, mse_list, model_name, best_correct_cent_list, best_mse_list, best_model_name_list):
    # choose best mean absolute error
    # correct_cent_list list
    # mse_list list
    # model_name string
    # best_correct_cent_list list
    # best_mse_list list
    # best_model_name_list list
    for i in range(len(correct_cent_list)):
        if correct_cent_list[i] < best_correct_cent_list[i]:
            best_correct_cent_list[i] = correct_cent_list[i]
            best_mse_list[i] = mse_list[i]
            best_model_name_list[i] = model_name
    return best_model_name_list, best_correct_cent_list, best_mse_list


def compare_game(game_0_list, game_1_list, game_2_list, game_3_list, model_name, best_game_0_list, best_game_1_list, best_game_2_list, best_game_3_list, best_model_name_list):
    # choose best grid average mean absolute error
    # correct_cent_list list
    # mse_list list
    # model_name string
    # best_correct_cent_list list
    # best_mse_list list
    # best_model_name_list list
    for i in range(len(game_0_list)):
        if game_0_list[i] < best_game_0_list[i]:
            best_game_0_list[i] = game_0_list[i]
            best_game_1_list[i] = game_1_list[i]
            best_game_2_list[i] = game_2_list[i]
            best_game_3_list[i] = game_3_list[i]
            best_model_name_list[i] = model_name
    return best_model_name_list, best_game_0_list, best_game_1_list, best_game_2_list, best_game_3_list


def compare_correct_cent(correct_cent_list, model_name, best_correct_cent_list, best_model_name_list):
    # choose higher correct cent
    # correct_cent_list list
    # model_name string
    # best_correct_cent_list list
    # best_model_name_list list
    for i in range(len(correct_cent_list)):
        if correct_cent_list[i] > best_correct_cent_list[i]:
            best_correct_cent_list[i] = correct_cent_list[i]
            best_model_name_list[i] = model_name
    return best_model_name_list, best_correct_cent_list


def gray_to_bgr(gray_image, mode='jet'):
    # gray_image 2-D ndarray
    # output is a 3 channel uint8 rgb ndarray
    getColorMap = pyplot.get_cmap(mode)

    # gray_image = gray_image / np.max(gray_image)

    rgba_image = getColorMap(gray_image)
    rgb_image = np.delete(rgba_image, 3, 2)
    rgb_image = 255 * rgb_image

    return cv2.cvtColor(rgb_image.astype(np.uint8), cv2.COLOR_RGB2BGR)


def make_path(path):
    if not isinstance(path, str):
        raise Exception('Path need to be a string.')
    if not os.path.exists(path):
        os.makedirs(path)


def get_foreground_mask(ground_truth_map):
    # ground_truth_map numpy.ndarray shape=(1, 1, h, w)
    mask = np.zeros_like(ground_truth_map)
    mask[ground_truth_map > 0] = 1.0
    return mask


def dilate_mask(mask, kernel_size, iterations=1, dtype=torch.float32, is_cuda=True):
    # masks type: torch.Tensor, shape is (1, 1, h, w)
    # kernel_size type: int or tuple of int
    if isinstance(kernel_size, int):
        kernel = np.ones((kernel_size, kernel_size))
    else:
        raise Exception('invalid kernel_size type')

    mask = mask.data.cpu().numpy()

    # reshape (1, 1, h, w) to (h, w)
    mask = mask.reshape(mask.shape[2], mask.shape[3])

    # dilate operation
    mask = cv2.dilate(mask, kernel, iterations=iterations)

    # reshape (h, w) to (1, 1, h, w)
    mask = mask.reshape(1, 1, mask.shape[0], mask.shape[1])

    mask = torch.from_numpy(mask).to(dtype)
    if is_cuda:
        mask = mask.cuda()

    return mask


def erode_mask(mask, kernel_size, iterations=1, dtype=torch.float32, is_cuda=True):
    # masks type: torch.Tensor, shape is (1, 1, h, w)
    # kernel_size type: int
    if isinstance(kernel_size, int):
        kernel = np.ones((kernel_size, kernel_size))
    else:
        raise Exception('invalid kernel_size type')

    mask = mask.data.cpu().numpy()

    # reshape (1, 1, h, w) to (h, w)
    mask = mask.reshape(mask.shape[2], mask.shape[3])

    # dilate operation
    mask = cv2.erode(mask, kernel, iterations=iterations)

    # reshape (h, w) to (1, 1, h, w)
    mask = mask.reshape(1, 1, mask.shape[0], mask.shape[1])

    mask = torch.from_numpy(mask).to(dtype)
    if is_cuda:
        mask = mask.cuda()

    return mask


def gaussian_kernel(shape=(15, 15), sigma=4):
    """
    2D gaussian mask - should give the same result as MATLAB's
    fspecial('gaussian',[shape],[sigma])
    """
    m, n = [(ss - 1.) / 2. for ss in shape]
    y,x = np.ogrid[-m:m+1, -n:n+1]
    h = np.exp(-(x ** 2 + y ** 2) / (2. * sigma ** 2))
    h[h < np.finfo(h.dtype).eps * h.max()] = 0
    sumh = h.sum()
    if sumh != 0:
        h /= sumh
    return h


class ExcelLog:
    def __init__(self, path):
        # path: str, path of the excel file
        # datasets: list, name of every sheet
        # keys: list, name of every column
        if not isinstance(path, str):
            raise Exception('path should be a string')

        self.path = path
        self.alphabet = '_ABCDEFGHIJKLMNOPQRSTUVWXYZ'

        excel_book = excel.Workbook()

        # for i in range(len(self.datasets)):
        #     excel_sheet = excel_book.create_sheet(self.datasets[i], i)
        #
        #     for j in range(len(self.keys)):
        #         excel_sheet[self.get_cell_name(j, 1)] = self.keys[j]

        excel_book.save(self.path)
        return

    def get_cell_name(self, column, row):
        if not (isinstance(column, int) or isinstance(row, int)):
            raise Exception('column and row should be integer')

        return self.alphabet[column] + str(row)

    def add_log(self, log_dict):
        # log_dict: dict, a dict contains several dicts, every dict contains information of one dataset
        if not isinstance(log_dict, dict):
            raise Exception('log_dict should be a dictionary')

        excel_book = excel.load_workbook(self.path)

        for dataset_name in log_dict:
            log = log_dict[dataset_name]
            try:
                excel_sheet = excel_book.get_sheet_by_name(dataset_name)
            except KeyError:
                excel_sheet = excel_book.create_sheet(dataset_name)
                column = 1
                for name in log:
                    excel_sheet[self.get_cell_name(column, 1)] = name
                    column += 1

            row = excel_sheet.max_row + 1

            column = 1
            for name in log:
                excel_sheet[self.get_cell_name(column, row)] = log[name]
                column += 1

        excel_book.save(self.path)
        return


def calculate_game(ground_truth, estimate, L=0):
    # grid average mean absolute error
    # ground_truth Tensor: shape=(1, 1, h, w)
    # estimate Tensor: same shape of ground_truth
    ground_truth = ndarray_to_tensor(ground_truth, is_cuda=True)
    estimate = ndarray_to_tensor(estimate, is_cuda=True)
    height = ground_truth.shape[2]
    width = ground_truth.shape[3]
    times = math.sqrt(math.pow(4, L))
    padding_height = int(math.ceil(height / times) * times - height)
    padding_width = int(math.ceil(width / times) * times - width)
    if padding_height != 0 or padding_width != 0:
        m = nn.ZeroPad2d((0, padding_width, 0, padding_height))
        ground_truth = m(ground_truth)
        estimate = m(estimate)
        height = ground_truth.shape[2]
        width = ground_truth.shape[3]
    m = nn.AdaptiveAvgPool2d(int(times))
    ground_truth = m(ground_truth) * (height / times) * (width / times)
    estimate = m(estimate) * (height / times) * (width / times)
    game = torch.sum(torch.abs(ground_truth - estimate))
    return game.item()


# def calculate_game_not_include_pad(ground_truth, estimate, L=0):
#     # grid average mean absolute error
#     # ground_truth Tensor: shape=(1, 1, h, w)
#     # estimate Tensor: same shape of ground_truth
#     ground_truth = ndarray_to_tensor(ground_truth, is_cuda=True)
#     estimate = ndarray_to_tensor(estimate, is_cuda=True)
#     height = ground_truth.shape[2]
#     width = ground_truth.shape[3]
#     times = math.sqrt(math.pow(4, L))
#     grid_height = int(math.ceil(height / times))
#     grid_width = int(math.ceil(width / times))
#     padding_height = int(math.ceil((grid_height * times - height) / 2))
#     padding_width = int(math.ceil((grid_width * times - width) / 2))
#     m = nn.AvgPool2d((grid_height, grid_width), stride=(grid_height, grid_width), padding=(padding_height, padding_width), count_include_pad=False)
#     ground_truth = m(ground_truth) * (height / times) * (width / times)
#     estimate = m(estimate) * (height / times) * (width / times)
#     game = torch.sum(torch.abs(ground_truth - estimate))
#     return game.item()


def print_red(a_string):
    print('\033[91m' + a_string + '\033[0m')


def build_block(x, size):
    # x shape=(1, c, h, w)
    height = x.shape[2]
    width = x.shape[3]
    padding_height = math.ceil((math.ceil(height / size) * size - height) / 2)
    padding_width = math.ceil((math.ceil(width / size) * size - width) / 2)
    return functional.avg_pool2d(x, size, stride=size, padding=(padding_height, padding_width), count_include_pad=True) * size * size